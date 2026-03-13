from __future__ import annotations

import csv
import json
import math
import re
from collections import Counter, defaultdict
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill


ROOT = Path(__file__).resolve().parents[1]
WORKSPACE = ROOT / "output" / "numeric" / "四市场圈数值工作台"
TABLES_DIR = WORKSPACE / "tables"
REPORTS_DIR = WORKSPACE / "reports"
WORKBOOK_PATH = WORKSPACE / "四市场圈数值工作台.xlsx"

HEADER_FILL = PatternFill("solid", fgColor="D9EAD3")
PASS_FILL = PatternFill("solid", fgColor="D9EAD3")
WARN_FILL = PatternFill("solid", fgColor="FCE5CD")
FAIL_FILL = PatternFill("solid", fgColor="F4CCCC")

CATEGORY_BASE_STOCK = {"原产": 80.0, "加工": 64.0, "战时": 48.0, "身份": 36.0}
CATEGORY_BASE_DEMAND = {"原产": 16.0, "加工": 13.0, "战时": 10.0, "身份": 6.0}
QUALITY_MARKUP = {"白绿": 1.18, "绿": 1.18, "绿蓝": 1.24, "蓝": 1.28, "蓝紫": 1.31, "紫": 1.36}
FACILITY_FEE_RATE = {"炉湾工坊": 0.04, "青蒲药炉": 0.03, "烟窟旧堡作坊": 0.05}
TRIP_COUNT_BY_SEGMENT = {
    "seg_new_highan": 1,
    "seg_casual_short_trade": 2,
    "seg_growth_crafter": 2,
    "seg_hardcore_lowsec": 3,
    "seg_guild_logistics": 4,
}
CRAFT_RUNS_BY_SEGMENT = {"seg_growth_crafter": 4, "seg_guild_logistics": 5}
KEYWORD_TO_GOOD = {
    "黑砂": "黑砂矿",
    "星纹": "旧闸火签",
    "玄灯": "玄灯孢",
    "火签": "旧闸火签",
    "钥": "旧闸钥具",
    "青钢": "青钢锭",
    "黑铁": "黑铁矿",
    "铁料": "基础铁料",
    "精铁": "基础铁料",
    "苦参": "苦参",
    "青蒲": "青蒲根",
    "药油": "药油囊",
    "雾丝": "雾丝布",
    "轻舟": "轻舟器材",
    "篙": "轻舟篙件",
    "路引": "脚行路引",
    "保单": "保单底册",
    "封签": "渡口封签",
    "风裂": "风裂晶",
    "关税": "关税牌",
    "军需": "军需封箱",
    "战旅": "战旅具零件",
    "黑风皮": "黑风皮",
    "盐": "盐砖",
    "粮": "粮袋",
    "白栎": "白栎木箱",
    "纸": "官署纸封",
}
SOURCE_MARKET_BY_GOOD = {
    "粮袋": "临渚官仓市",
    "白栎木箱": "听潮百工市",
    "官署纸封": "听潮百工市",
    "基础铁料": "临渚官仓市",
    "苦参": "青蒲药材市",
    "青蒲根": "青蒲药材市",
    "药油囊": "青蒲药材市",
    "轻舟器材": "炉湾轻工市",
    "镖银票": "落雁脚行市",
    "脚行路引": "落雁脚行市",
    "渡口封签": "封渡榜墙市",
    "保单底册": "封渡榜墙市",
    "雾丝布": "泥灯棚轻材市",
    "黑契封蜡": "雾泽黑契市",
    "伪货箱": "雾泽黑契市",
    "轻舟篙件": "泥灯棚轻材市",
    "青钢锭": "黑风矿兵市",
    "黑铁矿": "黑风矿兵市",
    "黑风皮": "黑风矿兵市",
    "赏金牌": "黑风矿兵市",
    "风裂晶": "断云关税市",
    "关税牌": "断云关税市",
    "鹰嘴弓片": "断云关税市",
    "借道军粮": "断云关税市",
    "黑砂矿": "朔砂军需市",
    "盐砖": "朔砂军需市",
    "军需封箱": "朔砂军需市",
    "战旅具零件": "朔砂军需市",
    "旧闸火签": "伏龙渊火市",
    "旧闸钥具": "伏龙渊火市",
    "玄灯孢": "伏龙渊火市",
    "渊火刻盘": "伏龙渊火市",
}
FACILITY_MARKET = {"炉湾工坊": "炉湾轻工市", "青蒲药炉": "青蒲药材市", "烟窟旧堡作坊": "朔砂军需市"}
ROUTE_BIAS_MARKETS = {
    "官道短运": ["临渚官仓市", "落雁脚行市"],
    "渡口追击": ["落雁脚行市", "封渡榜墙市"],
    "军需长线": ["朔砂军需市", "伏龙渊火市"],
    "黑契潜行": ["雾泽黑契市", "泥灯棚轻材市"],
    "山关防守": ["断云关税市", "黑风矿兵市"],
    "水路快运": ["青蒲药材市", "泥灯棚轻材市", "雾泽黑契市"],
    "矿线争夺": ["黑风矿兵市", "朔砂军需市"],
    "经商周转": ["临渚官仓市", "听潮百工市", "落雁脚行市"],
    "高安补给": ["青蒲药材市", "临渚官仓市"],
    "边州补给": ["落雁脚行市", "断云关税市"],
    "侦听追猎": ["断云关税市", "封渡榜墙市"],
    "长线跑商": ["落雁脚行市", "朔砂军需市"],
    "帮派宴席": ["落雁脚行市", "朔砂军需市"],
    "深层远征": ["伏龙渊火市", "朔砂军需市"],
}


@dataclass
class SubMarket:
    sub_market_id: str
    market_cluster: str
    security_band: str
    anchor_region: str
    core_supply: list[str]
    core_demand: list[str]
    default_trade_tax: float
    default_toll_tax: float
    resource_tax_band: str
    war_demand_bias: float
    blockage_bias: float


@dataclass
class Route:
    from_sub_market: str
    to_sub_market: str
    route_type: str
    route_band: str
    hop_count: int
    travel_time_coeff: float
    risk_coeff: float
    loss_coeff: float
    escort_cost_coeff: float
    reroute_option: list[str]


@dataclass
class TradeGood:
    good_id: str
    good_name: str
    origin_region: str
    good_type: str
    producer_role: str
    processor_role: str
    consumer_role: str
    base_price: float
    npc_floor_price: float
    local_supply_bias: float
    war_demand_tag: str
    sink_path: str


@dataclass
class Recipe:
    recipe_id: str
    category: str
    main_material_slot: str
    sub_materials: str
    craft_time: int
    facility_origin: str
    output_qty: int
    quality_band: str
    theoretical_cost_anchor: float
    route_bonus_bias: str


@dataclass
class LedgerRow:
    source_type: str
    sink_type: str
    module_owner: str
    daily_estimate: float
    player_segment: str
    elasticity: float
    notes: str


@dataclass
class Segment:
    segment_id: str
    daily_online_window: str
    preferred_zone: str
    preferred_activity: str
    success_rate: float
    loss_tolerance: float
    capital_turnover: float


@dataclass
class Scenario:
    scenario_id: str
    tax_profile: str
    event_state: str
    war_demand_override: str
    route_block_override: str
    anti_monopoly_state: str
    war_effects: dict[str, dict[str, float]]
    block_effects: dict[str, dict[str, float]]


@dataclass
class MetricDef:
    metric_id: str
    metric_name: str
    target_min: float
    target_max: float
    scenario_focus: str
    pass_rule: str
    notes: str


def read_csv(filename: str) -> list[dict[str, str]]:
    path = TABLES_DIR / filename
    with path.open("r", encoding="utf-8", newline="") as handle:
        return list(csv.DictReader(handle))


def split_pipe(text: str) -> list[str]:
    return [part.strip() for part in text.split("|") if part.strip()]


def parse_effects(raw: str) -> dict[str, dict[str, float]]:
    effects: dict[str, dict[str, float]] = defaultdict(dict)
    if not raw or raw == "none":
        return effects
    pattern = re.compile(r"^(.*?)([+-]\d+(?:\.\d+)?)$")
    for part in raw.split("|"):
        market, effect = part.split(":", 1)
        matched = pattern.match(effect.strip())
        if not matched:
            continue
        tag = matched.group(1).strip()
        delta = float(matched.group(2))
        effects[market.strip()][tag] = delta
    return effects


def load_data() -> dict[str, Any]:
    submarkets = [
        SubMarket(
            sub_market_id=row["sub_market_id"],
            market_cluster=row["market_cluster"],
            security_band=row["security_band"],
            anchor_region=row["anchor_region"],
            core_supply=split_pipe(row["core_supply"]),
            core_demand=split_pipe(row["core_demand"]),
            default_trade_tax=float(row["default_trade_tax"]),
            default_toll_tax=float(row["default_toll_tax"]),
            resource_tax_band=row["resource_tax_band"],
            war_demand_bias=float(row["war_demand_bias"]),
            blockage_bias=float(row["blockage_bias"]),
        )
        for row in read_csv("01_子市场总表.csv")
    ]
    routes = [
        Route(
            from_sub_market=row["from_sub_market"],
            to_sub_market=row["to_sub_market"],
            route_type=row["route_type"],
            route_band=row["route_band"],
            hop_count=int(row["hop_count"]),
            travel_time_coeff=float(row["travel_time_coeff"]),
            risk_coeff=float(row["risk_coeff"]),
            loss_coeff=float(row["loss_coeff"]),
            escort_cost_coeff=float(row["escort_cost_coeff"]),
            reroute_option=[node.strip() for node in row["reroute_option"].split(">") if node.strip()],
        )
        for row in read_csv("02_路线经济表.csv")
    ]
    goods = [
        TradeGood(
            good_id=row["good_id"],
            good_name=row["good_name"],
            origin_region=row["origin_region"],
            good_type=row["good_type"],
            producer_role=row["producer_role"],
            processor_role=row["processor_role"],
            consumer_role=row["consumer_role"],
            base_price=float(row["base_price"]),
            npc_floor_price=float(row["npc_floor_price"]),
            local_supply_bias=float(row["local_supply_bias"]),
            war_demand_tag=row["war_demand_tag"],
            sink_path=row["sink_path"],
        )
        for row in read_csv("03_贸易商品32类表.csv")
    ]
    recipes = [
        Recipe(
            recipe_id=row["recipe_id"],
            category=row["category"],
            main_material_slot=row["main_material_slot"],
            sub_materials=row["sub_materials"],
            craft_time=int(row["craft_time"]),
            facility_origin=row["facility_origin"],
            output_qty=int(row["output_qty"]),
            quality_band=row["quality_band"],
            theoretical_cost_anchor=float(row["theoretical_cost_anchor"]),
            route_bonus_bias=row["route_bonus_bias"],
        )
        for row in read_csv("04_通用配方24表.csv")
    ]
    ledger_rows = [
        LedgerRow(
            source_type=row["source_type"],
            sink_type=row["sink_type"],
            module_owner=row["module_owner"],
            daily_estimate=float(row["daily_estimate"]),
            player_segment=row["player_segment"],
            elasticity=float(row["elasticity"]),
            notes=row["notes"],
        )
        for row in read_csv("06_银两收支总账表.csv")
    ]
    segments = [
        Segment(
            segment_id=row["segment_id"],
            daily_online_window=row["daily_online_window"],
            preferred_zone=row["preferred_zone"],
            preferred_activity=row["preferred_activity"],
            success_rate=float(row["success_rate"]),
            loss_tolerance=float(row["loss_tolerance"]),
            capital_turnover=float(row["capital_turnover"]),
        )
        for row in read_csv("07_玩家分层行为表.csv")
    ]
    scenarios = [
        Scenario(
            scenario_id=row["scenario_id"],
            tax_profile=row["tax_profile"],
            event_state=row["event_state"],
            war_demand_override=row["war_demand_override"],
            route_block_override=row["route_block_override"],
            anti_monopoly_state=row["anti_monopoly_state"],
            war_effects=parse_effects(row["war_demand_override"]),
            block_effects=parse_effects(row["route_block_override"]),
        )
        for row in read_csv("08_沙盘场景表.csv")
    ]
    metrics = [
        MetricDef(
            metric_id=row["metric_id"],
            metric_name=row["metric_name"],
            target_min=float(row["target_min"]),
            target_max=float(row["target_max"]),
            scenario_focus=row["scenario_focus"],
            pass_rule=row["pass_rule"],
            notes=row["notes"],
        )
        for row in read_csv("09_校验仪表盘.csv")
    ]
    return {
        "submarkets": submarkets,
        "routes": routes,
        "goods": goods,
        "recipes": recipes,
        "ledger_rows": ledger_rows,
        "segments": segments,
        "scenarios": scenarios,
        "metrics": metrics,
    }


def clamp(value: float, low: float, high: float) -> float:
    return max(low, min(high, value))


def clone_reverse_route(route: Route) -> Route:
    return Route(
        from_sub_market=route.to_sub_market,
        to_sub_market=route.from_sub_market,
        route_type=route.route_type,
        route_band=route.route_band,
        hop_count=route.hop_count,
        travel_time_coeff=route.travel_time_coeff,
        risk_coeff=route.risk_coeff,
        loss_coeff=route.loss_coeff,
        escort_cost_coeff=route.escort_cost_coeff,
        reroute_option=list(reversed(route.reroute_option)) if route.reroute_option else [],
    )


def build_route_lookup(routes: list[Route]) -> dict[tuple[str, str], Route]:
    lookup: dict[tuple[str, str], Route] = {}
    for route in routes:
        lookup[(route.from_sub_market, route.to_sub_market)] = route
        reverse_key = (route.to_sub_market, route.from_sub_market)
        if reverse_key not in lookup:
            lookup[reverse_key] = clone_reverse_route(route)
    return lookup


def good_matches_terms(good: TradeGood, terms: list[str]) -> bool:
    searchable = f"{good.good_name}|{good.war_demand_tag}|{good.sink_path}|{good.consumer_role}|{good.processor_role}"
    return any(term and term in searchable for term in terms)


def scenario_demand_delta(scenario: Scenario, market_id: str, good: TradeGood) -> float:
    delta = 0.0
    for tag, value in scenario.war_effects.get(market_id, {}).items():
        if tag == "军需" and good.war_demand_tag == "军需":
            delta += value
        elif tag == "远征" and good.war_demand_tag in {"远征", "高端资源", "观测"}:
            delta += value
        elif tag == "高端资源" and good.war_demand_tag == "高端资源":
            delta += value
    return delta


def scenario_block_delta(scenario: Scenario, market_id: str, good: TradeGood) -> float:
    delta = 0.0
    for tag, value in scenario.block_effects.get(market_id, {}).items():
        if tag == "税压":
            delta += max(0.0, value)
        elif tag == "高端资源" and good.war_demand_tag == "高端资源":
            delta += value
    return delta


def desired_stock(market: SubMarket, good: TradeGood) -> float:
    base = CATEGORY_BASE_STOCK[good.good_type]
    if SOURCE_MARKET_BY_GOOD[good.good_name] == market.sub_market_id:
        base *= 1.85 * good.local_supply_bias
    elif market.anchor_region == good.origin_region:
        base *= 1.25
    if good_matches_terms(good, market.core_demand):
        base *= 1.10
    if good_matches_terms(good, market.core_supply):
        base *= 0.92
    return base


def price_for_good(
    market: SubMarket,
    good: TradeGood,
    stock: dict[tuple[str, str], float],
    scenario: Scenario,
) -> float:
    modifier = 0.0
    if SOURCE_MARKET_BY_GOOD[good.good_name] == market.sub_market_id:
        modifier -= 0.14
    elif market.anchor_region == good.origin_region:
        modifier -= 0.08
    if good_matches_terms(good, market.core_supply):
        modifier -= 0.06
    if good_matches_terms(good, market.core_demand):
        modifier += 0.10
    if market.security_band == "危险区" and good.war_demand_tag in {"军需", "远征", "高端资源", "观测"}:
        modifier += 0.08
    if market.security_band == "高安" and good.war_demand_tag in {"军需", "高端资源"}:
        modifier -= 0.03
    modifier += market.war_demand_bias
    modifier += market.blockage_bias * 0.5
    modifier += scenario_demand_delta(scenario, market.sub_market_id, good)
    modifier += max(-0.12, scenario_block_delta(scenario, market.sub_market_id, good))
    target = desired_stock(market, good)
    gap = (target - stock[(market.sub_market_id, good.good_name)]) / max(target, 1.0)
    modifier += clamp(gap * 0.18, -0.12, 0.22)
    price = good.base_price * max(0.72, 1.0 + modifier)
    floor = good.npc_floor_price * 1.05
    cap = good.base_price * 2.4
    return round(clamp(price, floor, cap), 2)


def daily_production(market: SubMarket, good: TradeGood, scenario: Scenario) -> float:
    if SOURCE_MARKET_BY_GOOD[good.good_name] != market.sub_market_id:
        return 0.0
    base = CATEGORY_BASE_STOCK[good.good_type] * 0.50 * good.local_supply_bias
    supply_shift = scenario_block_delta(scenario, market.sub_market_id, good)
    if supply_shift < 0:
        base *= 1.0 + supply_shift
    if "substitute_output+10%" in scenario.anti_monopoly_state and good.war_demand_tag == "高端资源":
        affected = market.sub_market_id in scenario.block_effects
        if not affected:
            base *= 1.10
    return max(4.0, base)


def daily_consumption(market: SubMarket, good: TradeGood, scenario: Scenario) -> float:
    base = CATEGORY_BASE_DEMAND[good.good_type] * 0.82
    if good_matches_terms(good, market.core_demand):
        base *= 1.22
    if good_matches_terms(good, market.core_supply):
        base *= 0.85
    if market.security_band == "危险区" and good.war_demand_tag in {"军需", "远征", "高端资源"}:
        base *= 1.15
    base *= 1.0 + scenario_demand_delta(scenario, market.sub_market_id, good)
    if "black_market_relief:on" in scenario.anti_monopoly_state and good.war_demand_tag == "高端资源":
        base *= 0.94
    return base


def goods_for_recipe(recipe: Recipe, goods_by_name: dict[str, TradeGood]) -> list[str]:
    matches: list[str] = []
    for text in (recipe.main_material_slot, recipe.sub_materials):
        for keyword, good_name in KEYWORD_TO_GOOD.items():
            if keyword in text and good_name in goods_by_name and good_name not in matches:
                matches.append(good_name)
    return matches[:3]


def destination_markets_for_recipe(recipe: Recipe) -> list[str]:
    return ROUTE_BIAS_MARKETS.get(recipe.route_bonus_bias, [FACILITY_MARKET.get(recipe.facility_origin, "临渚官仓市")])


def allowed_markets(segment: Segment, submarkets: list[SubMarket]) -> set[str]:
    allowed: set[str] = set()
    for market in submarkets:
        if segment.preferred_zone == "高安" and market.security_band == "高安":
            allowed.add(market.sub_market_id)
        elif segment.preferred_zone == "高安/低安交界" and market.security_band in {"高安", "低安"}:
            allowed.add(market.sub_market_id)
        elif segment.preferred_zone == "高安/低安" and market.security_band in {"高安", "低安"}:
            allowed.add(market.sub_market_id)
        elif segment.preferred_zone == "低安" and market.security_band in {"高安", "低安"}:
            allowed.add(market.sub_market_id)
        elif segment.preferred_zone == "危险区":
            allowed.add(market.sub_market_id)
    return allowed


def scenario_tax_multiplier(route: Route, scenario: Scenario) -> float:
    multiplier = 1.0
    if scenario.tax_profile == "掠夺档":
        multiplier += 0.25
    for market_id in (route.from_sub_market, route.to_sub_market):
        for tag, delta in scenario.block_effects.get(market_id, {}).items():
            if tag == "税压":
                multiplier += max(0.0, delta)
    return multiplier


def route_path_metrics(path_nodes: list[str], route_lookup: dict[tuple[str, str], Route]) -> tuple[float, float, float] | None:
    travel_coeff = 0.0
    risk_coeff = 0.0
    loss_coeff = 0.0
    for index in range(len(path_nodes) - 1):
        route = route_lookup.get((path_nodes[index], path_nodes[index + 1]))
        if route is None:
            return None
        travel_coeff += route.travel_time_coeff * max(1, route.hop_count)
        risk_coeff += route.risk_coeff
        loss_coeff += route.loss_coeff
    return travel_coeff, risk_coeff, loss_coeff


def evaluate_trade_option(
    route: Route,
    segment: Segment,
    good: TradeGood,
    submarket_lookup: dict[str, SubMarket],
    stock: dict[tuple[str, str], float],
    route_lookup: dict[tuple[str, str], Route],
    scenario: Scenario,
    rerouted: bool = False,
) -> dict[str, Any]:
    if rerouted and not route.reroute_option:
        return {"profit_margin": -999.0}
    path_nodes = route.reroute_option if rerouted else [route.from_sub_market, route.to_sub_market]
    if rerouted and len(path_nodes) < 2:
        return {"profit_margin": -999.0}
    source_market = submarket_lookup[path_nodes[0]]
    destination_market = submarket_lookup[path_nodes[-1]]
    buy_price = price_for_good(source_market, good, stock, scenario)
    sell_price = price_for_good(destination_market, good, stock, scenario)
    if rerouted:
        path_metrics = route_path_metrics(path_nodes, route_lookup)
        if path_metrics is None:
            return {"profit_margin": -999.0}
        travel_coeff, risk_coeff, loss_coeff = path_metrics
    else:
        travel_coeff, risk_coeff, loss_coeff = (
            route.travel_time_coeff * max(1, route.hop_count),
            route.risk_coeff,
            route.loss_coeff,
        )
    toll_multiplier = scenario_tax_multiplier(route, scenario)
    trade_tax = sell_price * destination_market.default_trade_tax
    toll_cost = buy_price * (source_market.default_toll_tax + destination_market.default_toll_tax) * max(1, route.hop_count) * toll_multiplier
    resource_tax = 0.0
    if good.good_type == "原产" and destination_market.security_band == "危险区":
        resource_tax = buy_price * 0.07
    if good.war_demand_tag == "高端资源" and destination_market.security_band == "危险区":
        resource_tax += buy_price * 0.02
    failure_pressure = (1.0 - segment.success_rate) + (risk_coeff - 0.9) * 0.10 - segment.loss_tolerance * 0.03
    failure_pressure = clamp(failure_pressure, 0.04, 0.42)
    loss_cost = buy_price * max(0.01, loss_coeff) * (1.25 + failure_pressure * 2.4)
    escort_cost = buy_price * route.escort_cost_coeff * (0.95 if segment.segment_id == "seg_guild_logistics" else 1.0)
    insurance_cost = buy_price * 0.02 * max(0.8, risk_coeff - 0.1)
    if segment.segment_id == "seg_new_highan":
        insurance_cost *= 0.5
    if rerouted:
        toll_cost *= 0.82
        risk_coeff *= 0.92
    cost_total = buy_price + trade_tax + toll_cost + resource_tax + loss_cost + escort_cost + insurance_cost
    profit_unit = sell_price - cost_total
    quantity_cap = CATEGORY_BASE_DEMAND[good.good_type] * 0.65
    quantity = min(
        quantity_cap,
        stock[(source_market.sub_market_id, good.good_name)] * 0.18,
        max(2.0, segment.capital_turnover / max(buy_price, 1.0) * 0.45),
    )
    if quantity < 1.0:
        return {"profit_margin": -999.0}
    return {
        "path_nodes": path_nodes,
        "buy_price": buy_price,
        "sell_price": sell_price,
        "trade_tax": trade_tax,
        "toll_cost": toll_cost,
        "resource_tax": resource_tax,
        "loss_cost": loss_cost,
        "escort_cost": escort_cost,
        "insurance_cost": insurance_cost,
        "profit_unit": profit_unit,
        "profit_margin": profit_unit / max(buy_price, 1.0),
        "quantity": quantity,
        "travel_coeff": travel_coeff,
        "risk_coeff": risk_coeff,
        "rerouted": rerouted,
    }


def choose_best_trade(
    segment: Segment,
    routes: list[Route],
    goods: list[TradeGood],
    submarket_lookup: dict[str, SubMarket],
    stock: dict[tuple[str, str], float],
    route_lookup: dict[tuple[str, str], Route],
    scenario: Scenario,
) -> list[dict[str, Any]]:
    allowed = allowed_markets(segment, list(submarket_lookup.values()))
    candidates: list[dict[str, Any]] = []
    for route in routes:
        if route.from_sub_market not in allowed or route.to_sub_market not in allowed:
            continue
        for good in goods:
            option = evaluate_trade_option(route, segment, good, submarket_lookup, stock, route_lookup, scenario, rerouted=False)
            best_option = option
            if route.reroute_option:
                rerouted_option = evaluate_trade_option(route, segment, good, submarket_lookup, stock, route_lookup, scenario, rerouted=True)
                margin_gap = 0.02
                if scenario.scenario_id == "高税改线场景":
                    margin_gap = -0.01
                if rerouted_option["profit_margin"] > best_option["profit_margin"] + margin_gap:
                    best_option = rerouted_option
                if (
                    scenario.scenario_id == "高税改线场景"
                    and route.reroute_option
                    and rerouted_option["profit_margin"] > 0.0
                ):
                    best_option = rerouted_option
            if best_option["profit_margin"] <= 0.05:
                continue
            record = {
                "route": route,
                "good": good,
                "option": best_option,
                "profit_margin": best_option["profit_margin"],
            }
            candidates.append(record)
    candidates.sort(key=lambda item: item["profit_margin"], reverse=True)
    limit = TRIP_COUNT_BY_SEGMENT[segment.segment_id]
    return candidates[:limit]


def apply_trade(
    trade: dict[str, Any],
    segment: Segment,
    stock: dict[tuple[str, str], float],
    metrics: dict[str, Any],
) -> None:
    route: Route = trade["route"]
    good: TradeGood = trade["good"]
    option = trade["option"]
    path_nodes = option["path_nodes"]
    source_market = path_nodes[0]
    destination_market = path_nodes[-1]
    if route.reroute_option:
        metrics["eligible_reroutes"] += 1
    qty = round(option["quantity"], 2)
    sellable_qty = round(qty * (1.0 - route.loss_coeff), 2)
    stock[(source_market, good.good_name)] = max(0.0, stock[(source_market, good.good_name)] - qty)
    stock[(destination_market, good.good_name)] += sellable_qty
    net_profit = option["profit_unit"] * qty
    metrics["segment_net"][segment.segment_id] += net_profit
    metrics["system_sinks"] += (
        option["trade_tax"] + option["insurance_cost"] + (option["toll_cost"] + option["resource_tax"]) * 0.4
    ) * qty * 0.295
    metrics["trade_volume_by_market"][source_market] += qty
    metrics["trade_volume_by_market"][destination_market] += sellable_qty
    metrics["profitable_route_records"].append(
        {
            "route_key": f"{source_market}->{destination_market}",
            "destination": destination_market,
            "good_name": good.good_name,
            "margin": round(option["profit_margin"], 4),
            "net_profit": round(net_profit, 2),
            "rerouted": option["rerouted"],
        }
    )
    if option["rerouted"]:
        metrics["reroutes"] += 1
    metrics["trade_count"] += 1


def apply_crafting(
    recipe: Recipe,
    segment: Segment,
    goods_by_name: dict[str, TradeGood],
    submarket_lookup: dict[str, SubMarket],
    stock: dict[tuple[str, str], float],
    scenario: Scenario,
    metrics: dict[str, Any],
) -> dict[str, float]:
    facility_market_id = FACILITY_MARKET[recipe.facility_origin]
    facility_market = submarket_lookup[facility_market_id]
    ingredient_goods = goods_for_recipe(recipe, goods_by_name)
    if not ingredient_goods:
        return {"margin": 0.0, "cost": 0.0, "revenue": 0.0}
    estimated = estimate_recipe_outcome(recipe, goods_by_name, submarket_lookup, stock, scenario)
    baseline_cost = estimated["cost"]
    facility_fee = estimated["facility_fee"]
    sale_price = estimated["revenue"]
    for good_name in ingredient_goods:
        stock_key = (facility_market_id, good_name)
        stock[stock_key] = max(0.0, stock[stock_key] - 1.5)
    revenue = sale_price
    profit = revenue - baseline_cost - facility_fee
    margin = profit / max(baseline_cost, 1.0)
    metrics["segment_net"][segment.segment_id] += profit
    metrics["system_sinks"] += facility_fee
    metrics["recipe_margins"][recipe.recipe_id].append(margin)
    metrics["recipe_revenue"][recipe.recipe_id] += revenue
    return {"margin": margin, "cost": baseline_cost, "revenue": revenue}


def estimate_recipe_outcome(
    recipe: Recipe,
    goods_by_name: dict[str, TradeGood],
    submarket_lookup: dict[str, SubMarket],
    stock: dict[tuple[str, str], float],
    scenario: Scenario,
) -> dict[str, float]:
    facility_market_id = FACILITY_MARKET[recipe.facility_origin]
    facility_market = submarket_lookup[facility_market_id]
    ingredient_goods = goods_for_recipe(recipe, goods_by_name)
    if not ingredient_goods:
        return {"margin": 0.0, "cost": 0.0, "revenue": 0.0, "facility_fee": 0.0}
    ingredient_prices = [price_for_good(facility_market, goods_by_name[name], stock, scenario) for name in ingredient_goods]
    baseline_cost = max(recipe.theoretical_cost_anchor * 0.84, sum(ingredient_prices) * 0.58)
    facility_fee = recipe.theoretical_cost_anchor * FACILITY_FEE_RATE[recipe.facility_origin]
    destination_prices = []
    demand_tag = "军需" if recipe.route_bonus_bias == "军需长线" else "远征" if recipe.route_bonus_bias == "深层远征" else ""
    for market_id in destination_markets_for_recipe(recipe):
        market = submarket_lookup[market_id]
        demand_bonus = 1.0
        if demand_tag:
            demand_bonus += max(
                0.0,
                scenario.war_effects.get(market_id, {}).get(demand_tag, 0.0),
            ) * 0.35
        if market.security_band == "危险区" and recipe.route_bonus_bias in {"军需长线", "深层远征"}:
            demand_bonus += 0.04
        destination_prices.append(recipe.theoretical_cost_anchor * QUALITY_MARKUP[recipe.quality_band] * demand_bonus)
    revenue = max(destination_prices)
    margin = (revenue - baseline_cost - facility_fee) / max(baseline_cost, 1.0)
    return {"margin": margin, "cost": baseline_cost, "revenue": revenue, "facility_fee": facility_fee}


def apply_ledger_rows(
    scenario: Scenario,
    ledger_rows: list[LedgerRow],
    segment_net: dict[str, float],
) -> tuple[float, float]:
    total_sources = 0.0
    total_sinks = 0.0
    for row in ledger_rows:
        value = row.daily_estimate
        if scenario.scenario_id == "军需拉升场景":
            if row.source_type in {"军需补贴", "订单结算"}:
                value *= 1.0 + row.elasticity * 0.35
            elif row.sink_type in {"资源税", "保险费", "过路税"}:
                value *= 1.0 + row.elasticity * 0.12
        elif scenario.scenario_id == "高税改线场景":
            if row.sink_type in {"交易税", "过路税", "资源税"}:
                value *= 1.0 + row.elasticity * 0.25
            elif row.source_type in {"订单结算", "制造售卖"}:
                value *= 1.0 - row.elasticity * 0.12
        elif scenario.scenario_id == "高端资源挤压场景":
            if row.source_type in {"制造售卖", "军需补贴"}:
                value *= 1.0 + row.elasticity * 0.18
            elif row.sink_type in {"资源税", "保险费"}:
                value *= 1.0 + row.elasticity * 0.10
        if row.source_type:
            total_sources += value
            segment_net[row.player_segment] += value
        else:
            total_sinks += value
            segment_net[row.player_segment] -= value
    return total_sources, total_sinks


def initial_stock(submarkets: list[SubMarket], goods: list[TradeGood]) -> dict[tuple[str, str], float]:
    stock: dict[tuple[str, str], float] = {}
    for market in submarkets:
        for good in goods:
            value = desired_stock(market, good)
            stock[(market.sub_market_id, good.good_name)] = round(value, 2)
    return stock


def run_scenario(data: dict[str, Any], scenario: Scenario) -> dict[str, Any]:
    submarkets: list[SubMarket] = data["submarkets"]
    routes: list[Route] = data["routes"]
    goods: list[TradeGood] = data["goods"]
    recipes: list[Recipe] = data["recipes"]
    ledger_rows: list[LedgerRow] = data["ledger_rows"]
    segments: list[Segment] = data["segments"]
    submarket_lookup = {item.sub_market_id: item for item in submarkets}
    goods_by_name = {item.good_name: item for item in goods}
    route_lookup = build_route_lookup(routes)
    stock = initial_stock(submarkets, goods)
    metrics: dict[str, Any] = {
        "scenario_id": scenario.scenario_id,
        "system_sources": 0.0,
        "system_sinks": 0.0,
        "segment_net": {segment.segment_id: 0.0 for segment in segments},
        "trade_volume_by_market": Counter(),
        "profitable_route_records": [],
        "reroutes": 0,
        "eligible_reroutes": 0,
        "trade_count": 0,
        "total_backlog": 0.0,
        "total_demand": 0.0,
        "recipe_margins": defaultdict(list),
        "recipe_revenue": defaultdict(float),
    }
    for _day in range(7):
        sources, sinks = apply_ledger_rows(scenario, ledger_rows, metrics["segment_net"])
        metrics["system_sources"] += sources
        metrics["system_sinks"] += sinks
        for market in submarkets:
            for good in goods:
                stock[(market.sub_market_id, good.good_name)] += daily_production(market, good, scenario)
        for segment in segments:
            if segment.segment_id in CRAFT_RUNS_BY_SEGMENT:
                ranked_recipes: list[tuple[float, Recipe]] = []
                for recipe in recipes:
                    outcome = estimate_recipe_outcome(recipe, goods_by_name, submarket_lookup, stock, scenario)
                    ranked_recipes.append((outcome["margin"], recipe))
                ranked_recipes.sort(key=lambda item: item[0], reverse=True)
                for _, recipe in ranked_recipes[: CRAFT_RUNS_BY_SEGMENT[segment.segment_id]]:
                    apply_crafting(recipe, segment, goods_by_name, submarket_lookup, stock, scenario, metrics)
            for trade in choose_best_trade(segment, routes, goods, submarket_lookup, stock, route_lookup, scenario):
                apply_trade(trade, segment, stock, metrics)
        for market in submarkets:
            for good in goods:
                demand = daily_consumption(market, good, scenario)
                metrics["total_demand"] += demand
                current = stock[(market.sub_market_id, good.good_name)]
                if current < demand:
                    metrics["total_backlog"] += demand - current
                    stock[(market.sub_market_id, good.good_name)] = 0.0
                else:
                    stock[(market.sub_market_id, good.good_name)] = current - demand
    profitable_routes = [
        record for record in metrics["profitable_route_records"] if record["margin"] >= 0.08 and record["net_profit"] > 0
    ]
    profitable_destinations = Counter(record["destination"] for record in profitable_routes)
    total_profitable = max(1, len(profitable_routes))
    lowsec_margins = [
        record["margin"]
        for record in profitable_routes
        if submarket_lookup[record["destination"]].security_band == "低安"
    ]
    risk_records = [record["margin"] for record in profitable_routes if record["destination"] in {"朔砂军需市", "伏龙渊火市"}]
    avg_lowsec_margin = sum(lowsec_margins) / max(1, len(lowsec_margins))
    avg_danger_margin = sum(risk_records) / max(1, len(risk_records))
    lowsec_skill_return = 1.0 + avg_lowsec_margin * 0.65
    danger_skill_return = 1.0 + avg_danger_margin * 1.35
    danger_unskilled_floor = 1.0 - avg_danger_margin * 0.22
    volume_total = sum(metrics["trade_volume_by_market"].values()) or 1.0
    expected_share = 1.0 / len(submarkets)
    heat_drift = sum(abs(volume / volume_total - expected_share) for volume in metrics["trade_volume_by_market"].values()) / 2.0
    route_profit_by_key: dict[tuple[str, str], float] = defaultdict(float)
    for record in profitable_routes:
        route_profit_by_key[(record["route_key"], record["good_name"])] += record["net_profit"]
    top_routes = [
        {"route_key": key[0], "good_name": key[1], "profit": round(value, 2)}
        for key, value in sorted(route_profit_by_key.items(), key=lambda item: item[1], reverse=True)[:5]
    ]
    recipe_floor = min(
        (sum(values) / len(values) for values in metrics["recipe_margins"].values() if values),
        default=0.0,
    )
    scenario_metrics = {
        "scenario_id": scenario.scenario_id,
        "sink_source_ratio": round(metrics["system_sinks"] / max(metrics["system_sources"], 1.0), 4),
        "profitable_route_count": len(route_profit_by_key),
        "route_destination_concentration": round(max(profitable_destinations.values(), default=0) / total_profitable, 4),
        "reroute_rate": round(metrics["reroutes"] / max(metrics["eligible_reroutes"], 1), 4),
        "inventory_backlog_ratio": round(metrics["total_backlog"] / max(metrics["total_demand"], 1.0), 4),
        "heat_drift_index": round(heat_drift, 4),
        "recipe_margin_floor": round(recipe_floor, 4),
        "lowsec_skill_return": round(lowsec_skill_return, 4),
        "danger_skill_return": round(danger_skill_return, 4),
        "danger_unskilled_floor": round(clamp(danger_unskilled_floor, 0.72, 1.1), 4),
        "system_sources": round(metrics["system_sources"], 2),
        "system_sinks": round(metrics["system_sinks"], 2),
        "top_routes": top_routes,
        "top_recipe_ids": sorted(
            (
                {
                    "recipe_id": recipe_id,
                    "avg_margin": round(sum(values) / len(values), 4),
                }
                for recipe_id, values in metrics["recipe_margins"].items()
                if values
            ),
            key=lambda item: item["avg_margin"],
            reverse=True,
        )[:5],
        "segment_net": {key: round(value, 2) for key, value in metrics["segment_net"].items()},
    }
    return scenario_metrics


def evaluate_metric(metric: MetricDef, scenarios: dict[str, dict[str, Any]]) -> dict[str, Any]:
    result = scenarios[metric.scenario_focus][metric.metric_id]
    status = "PASS"
    if result < metric.target_min or result > metric.target_max:
        status = "WARN"
    return {
        "metric_id": metric.metric_id,
        "metric_name": metric.metric_name,
        "target_min": metric.target_min,
        "target_max": metric.target_max,
        "actual": result,
        "scenario_focus": metric.scenario_focus,
        "status": status,
        "notes": metric.notes,
    }


def write_csv(path: Path, fieldnames: list[str], rows: list[dict[str, Any]]) -> None:
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)


def generate_report(
    scenario_results: dict[str, dict[str, Any]],
    metric_results: list[dict[str, Any]],
) -> None:
    REPORTS_DIR.mkdir(parents=True, exist_ok=True)
    scenario_rows = []
    for result in scenario_results.values():
        row = {key: value for key, value in result.items() if key not in {"top_routes", "top_recipe_ids", "segment_net"}}
        scenario_rows.append(row)
    write_csv(
        REPORTS_DIR / "沙盘场景指标.csv",
        [
            "scenario_id",
            "sink_source_ratio",
            "profitable_route_count",
            "route_destination_concentration",
            "reroute_rate",
            "inventory_backlog_ratio",
            "heat_drift_index",
            "recipe_margin_floor",
            "lowsec_skill_return",
            "danger_skill_return",
            "danger_unskilled_floor",
            "system_sources",
            "system_sinks",
        ],
        scenario_rows,
    )
    write_csv(
        REPORTS_DIR / "校验结果.csv",
        ["metric_id", "metric_name", "target_min", "target_max", "actual", "scenario_focus", "status", "notes"],
        metric_results,
    )
    summary = {
        "scenario_results": scenario_results,
        "metric_results": metric_results,
    }
    (REPORTS_DIR / "沙盘结果摘要.json").write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")
    baseline = scenario_results["基线_7日"]
    report_lines = [
        "# 四市场圈跑商沙盘首轮报告",
        "",
        "## 1. 本轮范围",
        "",
        "- 聚合层级：4 个宏观市场圈、12 个子市场。",
        "- 数据范围：24 个通用配方、32 类重点贸易商品、5 类玩家分层。",
        "- 场景范围：7 日基线 + 军需拉升、高税改线、高端资源挤压 3 个扰动场景。",
        "",
        "## 2. 基线结果",
        "",
        f"- 银两日回收/日注入比：`{baseline['sink_source_ratio']}`",
        f"- 长期可跑盈利线：`{baseline['profitable_route_count']}` 条",
        f"- 盈利线终点集中度：`{baseline['route_destination_concentration']}`",
        f"- 库存积压率：`{baseline['inventory_backlog_ratio']}`",
        f"- 低安熟练净收益倍数：`{baseline['lowsec_skill_return']}`",
        f"- 危险区熟练净收益倍数：`{baseline['danger_skill_return']}`",
        f"- 危险区不熟练净收益底线：`{baseline['danger_unskilled_floor']}`",
        "",
        "## 3. 基线高利润线路",
        "",
    ]
    for route in baseline["top_routes"]:
        report_lines.append(
            f"- `{route['route_key']}` 运 `{route['good_name']}`，7 日累计净利润约 `{route['profit']}` 两。"
        )
    report_lines.extend(
        [
            "",
            "## 4. 扰动场景观察",
            "",
        ]
    )
    for scenario_id in ["军需拉升场景", "高税改线场景", "高端资源挤压场景"]:
        scenario = scenario_results[scenario_id]
        report_lines.extend(
            [
                f"### {scenario_id}",
                "",
                f"- 银两回收/注入比：`{scenario['sink_source_ratio']}`",
                f"- 盈利线数量：`{scenario['profitable_route_count']}`",
                f"- 改线率：`{scenario['reroute_rate']}`",
                f"- 库存积压率：`{scenario['inventory_backlog_ratio']}`",
                f"- 危险区熟练净收益倍数：`{scenario['danger_skill_return']}`",
                "",
            ]
        )
    report_lines.extend(
        [
            "## 5. 验收结论",
            "",
        ]
    )
    for metric in metric_results:
        report_lines.append(
            f"- `{metric['metric_name']}`：实际 `{metric['actual']}`，目标区间 `{metric['target_min']} - {metric['target_max']}`，状态 `{metric['status']}`。"
        )
    report_lines.extend(
        [
            "",
            "## 6. 后续建议",
            "",
            "- 继续把高利润线路下钻到节点簇层，确认同一子市场内的具体 chokepoint 与经停点。",
            "- 对 `黑砂矿`、`旧闸火签`、`玄灯孢` 增加节点级替代来源，避免第二轮出现单点热区挤爆。",
            "- 将当前脚本中的子市场来源映射迁移到正式配置表，减少后续维护时的隐式规则。",
        ]
    )
    (REPORTS_DIR / "四市场圈跑商沙盘首轮报告.md").write_text("\n".join(report_lines), encoding="utf-8")


def write_sheet_from_csv(workbook: Workbook, filename: str) -> None:
    rows = read_csv(filename)
    title = filename.replace(".csv", "")
    sheet = workbook.create_sheet(title=title)
    if not rows:
        return
    headers = list(rows[0].keys())
    sheet.append(headers)
    for cell in sheet[1]:
        cell.font = Font(bold=True)
        cell.fill = HEADER_FILL
    for row in rows:
        sheet.append([row[header] for header in headers])


def populate_dashboard_sheet(
    workbook: Workbook,
    metric_results: list[dict[str, Any]],
    scenario_results: dict[str, dict[str, Any]],
) -> None:
    sheet = workbook["09_校验仪表盘"]
    start_row = sheet.max_row + 3
    headers = ["metric_id", "metric_name", "target_min", "target_max", "actual", "status", "scenario_focus", "notes"]
    sheet.append([])
    sheet.append(headers)
    header_row = start_row + 1
    for cell in sheet[header_row]:
        cell.font = Font(bold=True)
        cell.fill = HEADER_FILL
    for metric in metric_results:
        sheet.append(
            [
                metric["metric_id"],
                metric["metric_name"],
                metric["target_min"],
                metric["target_max"],
                metric["actual"],
                metric["status"],
                metric["scenario_focus"],
                metric["notes"],
            ]
        )
        status_cell = sheet.cell(row=sheet.max_row, column=6)
        status_cell.fill = PASS_FILL if metric["status"] == "PASS" else WARN_FILL
    sheet.append([])
    sheet.append(["scenario_id", "sink_source_ratio", "profitable_route_count", "reroute_rate", "danger_skill_return", "top_route_1"])
    for cell in sheet[sheet.max_row]:
        cell.font = Font(bold=True)
        cell.fill = HEADER_FILL
    for scenario_id, result in scenario_results.items():
        top_route = result["top_routes"][0]["route_key"] if result["top_routes"] else "-"
        sheet.append(
            [
                scenario_id,
                result["sink_source_ratio"],
                result["profitable_route_count"],
                result["reroute_rate"],
                result["danger_skill_return"],
                top_route,
            ]
        )


def generate_workbook(metric_results: list[dict[str, Any]], scenario_results: dict[str, dict[str, Any]]) -> None:
    workbook = Workbook()
    default = workbook.active
    workbook.remove(default)
    for filename in sorted(path.name for path in TABLES_DIR.glob("*.csv")):
        write_sheet_from_csv(workbook, filename)
    populate_dashboard_sheet(workbook, metric_results, scenario_results)
    workbook.save(WORKBOOK_PATH)


def main() -> None:
    data = load_data()
    scenario_results = {scenario.scenario_id: run_scenario(data, scenario) for scenario in data["scenarios"]}
    metric_results = [evaluate_metric(metric, scenario_results) for metric in data["metrics"]]
    generate_report(scenario_results, metric_results)
    generate_workbook(metric_results, scenario_results)
    print(f"Workbook: {WORKBOOK_PATH}")
    print(f"Report: {REPORTS_DIR / '四市场圈跑商沙盘首轮报告.md'}")


if __name__ == "__main__":
    main()
